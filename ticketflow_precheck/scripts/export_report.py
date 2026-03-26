#!/usr/bin/env python3
"""
异常数据导出：将预检发现的异常生成 Excel 报告
"""

import os
import uuid

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
RED_FONT = Font(color="CC0000", bold=True)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)


def export_exception_report(
    ticket_code: str,
    exceptions: list,
    output_dir: str = "/tmp",
) -> str:
    """
    将预检异常数据导出为 Excel 报告。

    Args:
        ticket_code: 工单编号
        exceptions: precheck_huodan 返回的 exceptions 列表
        output_dir: 输出目录

    Returns:
        生成的 Excel 文件路径
    """
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(
        output_dir, f"precheck_exceptions_{ticket_code}_{uuid.uuid4().hex[:6]}.xlsx"
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "预检异常报告"

    headers = ["序号", "异常类型", "相关字段", "异常描述", "影响行数", "示例数据"]
    ws.append(headers)

    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    type_labels = {
        "missing_data": "数据缺失",
        "duplicate_conflict": "重复冲突",
        "invalid_price": "价格异常",
        "negative_stock": "库存异常",
    }

    for idx, exc in enumerate(exceptions, start=1):
        samples_text = ""
        for sample in exc.get("samples", []):
            parts = [f"{k}={v}" for k, v in sample.items() if k != "row"]
            samples_text += f"第{sample.get('row', '?')}行: {', '.join(parts)}\n"

        row = [
            idx,
            type_labels.get(exc["type"], exc["type"]),
            exc.get("field", ""),
            exc.get("detail", ""),
            len(exc.get("rows", [])),
            samples_text.strip(),
        ]
        ws.append(row)

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=idx + 1, column=col_idx)
            cell.font = RED_FONT

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 60

    total_affected = sum(len(exc.get("rows", [])) for exc in exceptions)
    summary_row = ws.max_row + 2
    ws.cell(row=summary_row, column=1, value="汇总")
    ws.cell(row=summary_row, column=1).font = Font(bold=True, size=12)
    ws.cell(row=summary_row, column=2, value=f"共 {len(exceptions)} 类异常")
    ws.cell(row=summary_row, column=5, value=f"共影响 {total_affected} 行")

    wb.save(output_path)
    wb.close()

    print(f"异常报告已生成: {output_path}")
    return output_path


def build_feedback_message(ticket_code: str, exceptions: list) -> str:
    """
    根据异常列表构建反馈消息文本。

    Args:
        ticket_code: 工单编号
        exceptions: precheck_huodan 返回的 exceptions 列表

    Returns:
        格式化的反馈消息
    """
    type_labels = {
        "missing_data": "货单数据缺失",
        "duplicate_conflict": "货单SKU异常",
        "invalid_price": "货单价格异常",
        "negative_stock": "货单库存异常",
    }

    issues = []
    for exc in exceptions:
        label = type_labels.get(exc["type"], exc["type"])
        detail = exc.get("detail", "")
        count = len(exc.get("rows", []))
        issues.append(f"{label}（{detail}，影响 {count} 行）")

    issues_text = "、".join(issues)
    return f"工单编号 {ticket_code}，{issues_text}，请修正后重新上传。"


if __name__ == "__main__":
    test_exceptions = [
        {
            "type": "missing_data",
            "field": "品类",
            "rows": [3, 5, 8],
            "detail": "品类字段存在缺失",
            "samples": [
                {"row": 3, "商品SPU": "SPU001", "商品SKU": "SKU001", "品类": None},
            ],
        },
        {
            "type": "invalid_price",
            "field": "奥莱价",
            "rows": [10],
            "detail": "奥莱价存在小于等于 0 的数据",
            "samples": [
                {"row": 10, "商品SPU": "SPU005", "商品SKU": "SKU005", "奥莱价": 0},
            ],
        },
    ]

    path = export_exception_report("T000117738002663038508819", test_exceptions)
    msg = build_feedback_message("T000117738002663038508819", test_exceptions)
    print(msg)
