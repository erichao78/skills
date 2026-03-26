#!/usr/bin/env python3
"""
Excel 文件工具：下载、解析、创建标准活动货单
"""

import os
import uuid
from typing import Optional

import requests
from openpyxl import Workbook, load_workbook

STANDARD_COLUMNS = [
    "商品名称",
    "商品SPU",
    "商品SKU",
    "吊牌价",
    "奥莱价",
    "库存",
    "品类",
    "颜色",
    "尺码",
    "季节",
    "性别",
]


def download_excel(url: str, save_dir: str = "/tmp") -> str:
    """
    从 URL 下载 Excel 文件到本地。

    Args:
        url: Excel 文件的 URL
        save_dir: 保存目录

    Returns:
        本地文件路径
    """
    os.makedirs(save_dir, exist_ok=True)

    filename = url.split("/")[-1].split("?")[0]
    if not filename.endswith((".xlsx", ".xls")):
        filename = f"{uuid.uuid4().hex[:8]}.xlsx"

    local_path = os.path.join(save_dir, filename)

    response = requests.get(url, timeout=60)
    response.raise_for_status()

    with open(local_path, "wb") as f:
        f.write(response.content)

    print(f"文件已下载: {local_path} ({len(response.content)} bytes)")
    return local_path


def get_sheet_columns(file_path: str) -> dict:
    """
    获取 Excel 文件中每个 sheet 的列名。

    Args:
        file_path: Excel 文件路径

    Returns:
        dict: {sheet_name: [column_names]}
    """
    wb = load_workbook(file_path, read_only=True)
    result = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(cell) if cell is not None else "" for cell in row]
            break
        result[sheet_name] = headers

    wb.close()
    return result


def read_excel_data(file_path: str, sheet_name: Optional[str] = None) -> list:
    """
    读取 Excel 数据为 dict 列表（以第一行为表头）。

    Args:
        file_path: Excel 文件路径
        sheet_name: sheet 名称，为 None 时读取第一个 sheet

    Returns:
        list[dict]: 每行数据为一个 dict
    """
    wb = load_workbook(file_path, read_only=True)

    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return []

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    data = []

    for row in rows[1:]:
        if all(cell is None for cell in row):
            continue
        row_dict = {}
        for i, cell in enumerate(row):
            if i < len(headers) and headers[i]:
                row_dict[headers[i]] = cell
        data.append(row_dict)

    return data


def create_standard_huodan(
    data: list,
    column_mapping: dict,
    output_dir: str = "/tmp",
) -> str:
    """
    根据列映射关系，将原始数据转换为标准活动货单 Excel 文件。

    Args:
        data: 原始数据 list[dict]
        column_mapping: 映射关系 dict，key 为标准列名，value 为原始列名（None 表示该列留空）
        output_dir: 输出目录

    Returns:
        生成的 Excel 文件路径
    """
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(
        output_dir, f"standard_huodan_{uuid.uuid4().hex[:8]}.xlsx"
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "活动货单"

    ws.append(STANDARD_COLUMNS)

    for row_data in data:
        row = []
        for col_name in STANDARD_COLUMNS:
            source_col = column_mapping.get(col_name)
            if source_col and source_col in row_data:
                row.append(row_data[source_col])
            else:
                row.append(None)
        ws.append(row)

    wb.save(output_path)
    wb.close()

    print(f"标准活动货单已生成: {output_path} ({len(data)} 行数据)")
    return output_path


if __name__ == "__main__":
    test_url = (
        "https://ossmp.shanshan-business.com/15f875e79e7af69053b5fe5ff925c2c0@0x0.xlsx"
    )
    try:
        path = download_excel(test_url)
        columns = get_sheet_columns(path)
        for sheet, cols in columns.items():
            print(f"Sheet: {sheet}, Columns: {cols}")
        data = read_excel_data(path)
        print(f"数据行数: {len(data)}")
    except Exception as e:
        print(f"测试失败: {e}")
