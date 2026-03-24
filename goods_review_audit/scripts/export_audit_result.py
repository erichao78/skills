#!/usr/bin/env python3
"""
审核结果导出

将商品审核结果生成 Excel 报告并拷贝到目标目录。
Excel 包含单个 Sheet（审核报告），采用透视表布局：每个商品一行，每个检查项一列。
通过显示"通过"（绿色），不通过显示具体原因（红色）。

输入数据结构 (all_goods_results):
[
    {
        "goodsSn": "R2401001",
        "name": "商品名称",
        "passed": True,
        "fail_reasons": [],
        "checks": [
            {"name": "基本信息完整性", "passed": True, "detail": ""},
            {"name": "价格合规性", "passed": True, "detail": ""},
            {"name": "敏感词检测", "passed": True, "detail": ""},
            {"name": "主图数量", "passed": False, "detail": "当前2张，要求至少3张"},
        ],
    },
]
"""

import os
from datetime import datetime
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _apply_header_style(ws, col_widths: List[int]) -> None:
    """Apply consistent header styling and column widths."""
    from openpyxl.utils import get_column_letter

    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER
    for i, width in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = width


def _build_merged_sheet(ws, all_goods_results: List[Dict[str, Any]]) -> None:
    """Build single-sheet report: one row per product, one column per check item."""
    ws.title = "审核报告"

    check_names: List[str] = []
    seen: set = set()
    for goods in all_goods_results:
        for check in goods.get("checks", []):
            if check["name"] not in seen:
                check_names.append(check["name"])
                seen.add(check["name"])

    headers = (
        ["序号", "SPU编码", "商品名称", "审核结果"] + check_names + ["不通过原因汇总"]
    )
    ws.append(headers)

    fixed_widths = [6, 18, 30, 12]
    check_widths = [18] * len(check_names)
    tail_widths = [50]
    _apply_header_style(ws, fixed_widths + check_widths + tail_widths)

    for idx, goods in enumerate(all_goods_results, 1):
        result = "通过" if goods["passed"] else "不通过"
        reasons = "；".join(goods["fail_reasons"]) if goods["fail_reasons"] else ""
        checks_map = {c["name"]: c for c in goods.get("checks", [])}

        row_data = [idx, goods["goodsSn"], goods["name"], result]
        for cn in check_names:
            c = checks_map.get(cn)
            if c is None:
                row_data.append("-")
            elif c["passed"]:
                row_data.append("通过")
            else:
                row_data.append(c.get("detail", "不通过"))
        row_data.append(reasons)

        ws.append(row_data)
        row_num = ws.max_row

        for cell in ws[row_num]:
            cell.border = _THIN_BORDER
        ws.cell(row=row_num, column=4).fill = (
            _PASS_FILL if goods["passed"] else _FAIL_FILL
        )

        for col_offset, cn in enumerate(check_names):
            col_idx = 5 + col_offset
            c = checks_map.get(cn)
            if c is not None:
                ws.cell(row=row_num, column=col_idx).fill = (
                    _PASS_FILL if c["passed"] else _FAIL_FILL
                )


def export_audit_result(
    all_goods_results: List[Dict[str, Any]],
    store_name: str,
    shop_name: str,
) -> str:
    """
    生成 Excel 审核报告。

    Args:
        all_goods_results: 审核结果列表（结构见模块文档）
        store_name: 门店名称
        shop_name: 店铺/品牌名称

    Returns:
        生成的 Excel 文件的完整路径（供调用方作为附件发送给用户）
    """
    wb = Workbook()

    _build_merged_sheet(wb.active, all_goods_results)

    date_str = datetime.now().strftime("%Y%m%d")
    filename = f"审核报告_{store_name}_{shop_name}_{date_str}.xlsx"

    local_path = os.path.join(os.getcwd(), filename)
    wb.save(local_path)

    return local_path


def generate_summary_text(
    store_name: str,
    shop_name: str,
    all_goods_results: List[Dict[str, Any]],
) -> str:
    """
    生成审核结果的文字总结。

    Args:
        store_name: 门店名称
        shop_name: 店铺/品牌名称
        all_goods_results: 审核结果列表

    Returns:
        格式化的文字总结
    """
    total = len(all_goods_results)
    passed = sum(1 for g in all_goods_results if g["passed"])
    failed = total - passed

    api_rejected = sum(
        1
        for g in all_goods_results
        if any(
            "后端" in r or "API" in r or "无共享库" in r
            for r in g.get("fail_reasons", [])
        )
    )

    lines = [
        "【审核完成】",
        "",
        f"门店：{store_name} | 店铺：{shop_name}",
        f"商品总数：{total} 款",
        "",
        f"✅ 审核通过：{passed} 款",
        f"❌ 审核不通过：{failed} 款",
    ]

    if api_rejected > 0:
        lines.append("")
        lines.append(
            f"⚠️ 其中 {api_rejected} 款商品本地检测通过但后端审核被拒，具体原因见 Excel 报告"
        )

    lines.append("")
    lines.append("详细审核报告见附件 Excel 文件。")

    return "\n".join(lines)


if __name__ == "__main__":
    sample_data = [
        {
            "goodsSn": "R2401001",
            "name": "AIR MAX 2024 运动鞋",
            "passed": True,
            "fail_reasons": [],
            "checks": [
                {"name": "基本信息完整性", "passed": True, "detail": ""},
                {"name": "价格合规性", "passed": True, "detail": ""},
                {"name": "主图数量", "passed": True, "detail": ""},
            ],
        },
        {
            "goodsSn": "R2401002",
            "name": "ULTRA BOOST 跑步鞋",
            "passed": False,
            "fail_reasons": [
                "主图数量不足（当前2张，要求至少3张）",
                "详情图数量不足（当前4张，要求至少5张）",
            ],
            "checks": [
                {"name": "基本信息完整性", "passed": True, "detail": ""},
                {"name": "价格合规性", "passed": True, "detail": ""},
                {"name": "主图数量", "passed": False, "detail": "当前2张，要求至少3张"},
                {
                    "name": "详情图数量",
                    "passed": False,
                    "detail": "当前4张，要求至少5张",
                },
            ],
        },
    ]

    path = export_audit_result(sample_data, "宁波店", "LE COQ SPORTIF")
    print(f"Excel 报告已生成：{path}")

    summary = generate_summary_text("宁波店", "LE COQ SPORTIF", sample_data)
    print(summary)
